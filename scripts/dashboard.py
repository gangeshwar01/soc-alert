# scripts/dashboard.py
import os
import io
import math
import joblib
import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from sklearn.metrics import confusion_matrix, classification_report, precision_recall_fscore_support
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

st.set_page_config(page_title="CICIDS2018 Dashboard (CSV + Excel)", layout="wide")

     
# CONFIG (edit paths if needed)
     
MODEL_PATH_DEFAULT = r"D:\RealTime_Alert_Analysis\model\cicids2018_rf_model_A.joblib"
SCALER_PATH_DEFAULT = r"D:\RealTime_Alert_Analysis\model\cicids2018_scaler_A.joblib"
DEFAULT_CHUNK_SIZE = 40000

     
# SESSION INIT
     
if "processed" not in st.session_state:
    st.session_state.processed = False
if "out_df" not in st.session_state:
    st.session_state.out_df = None
if "prob_matrix" not in st.session_state:
    st.session_state.prob_matrix = None
if "classes" not in st.session_state:
    st.session_state.classes = None
if "line_df" not in st.session_state:
    st.session_state.line_df = None

     
# HELPERS
     
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace("[^a-z0-9_]+", "_", regex=True)
        .str.replace("__+", "_", regex=True)
        .str.strip("_")
    )
    return df

def safe_to_numeric(df: pd.DataFrame) -> pd.DataFrame:
    df = df.apply(pd.to_numeric, errors="coerce")
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.fillna(df.median(numeric_only=True)).fillna(0)
    return df

def classify_risk(pred_label, iso_score, high_thresh=-0.2, med_thresh=0.0):
    pl = str(pred_label).lower()
    attack_keywords = [
        "attack", "dos", "ddos", "bruteforce", "botnet", "injection",
        "xss", "portscan", "sql", "infiltration", "malware"
    ]
    if any(k in pl for k in attack_keywords):
        return "HIGH"
    if iso_score < high_thresh:
        return "HIGH"
    if iso_score < med_thresh:
        return "MEDIUM"
    return "LOW"

# Confusion matrix plotting with readable numbers
def plot_confusion_matrix_readable(y_true, y_pred, classes, figsize=(10,8), cmap="Blues"):
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    fig, ax = plt.subplots(figsize=figsize)
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)

    ax.set_title("Confusion Matrix")
    ax.set_xticks(np.arange(len(classes)))
    ax.set_yticks(np.arange(len(classes)))

    # auto font sizing based on number of classes
    label_fontsize = 10 if len(classes) <= 20 else max(6, int(200 / len(classes)))
    tick_fontsize = label_fontsize
    ax.set_xticklabels(classes, rotation=45, ha="right", fontsize=tick_fontsize)
    ax.set_yticklabels(classes, fontsize=tick_fontsize)

    # grid lines
    ax.set_xticks(np.arange(-.5, cm.shape[1], 1), minor=True)
    ax.set_yticks(np.arange(-.5, cm.shape[0], 1), minor=True)
    ax.grid(which="minor", color="gray", linestyle='-', linewidth=0.3, alpha=0.3)
    ax.tick_params(which="minor", bottom=False, left=False)

    max_val = cm.max() if cm.size else 0
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            val = cm[i, j]
            norm_val = val / max_val if max_val > 0 else 0
            rgba = plt.cm.get_cmap(cmap)(norm_val)
            luminance = 0.299 * rgba[0] + 0.587 * rgba[1] + 0.114 * rgba[2]
            text_color = "black" if luminance > 0.5 else "white"
            ax.text(j, i, f"{val:,}", ha="center", va="center",
                    fontsize=max(7, int(label_fontsize * 0.8)),
                    color=text_color,
                    bbox=dict(facecolor="white", edgecolor="none", boxstyle="round,pad=0.2", alpha=0.85))
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    plt.tight_layout()
    return fig, cm

     
# LOAD MODEL (cache)
     
@st.cache_resource
def load_bundle(mpath, spath):
    bundle = joblib.load(mpath)
    clf = bundle.get("classifier")
    iso = bundle.get("isolation_forest", None)
    features = bundle.get("feature_names", []) or []
    label_enc = bundle.get("label_encoder", None)
    scaler = joblib.load(spath)
    return clf, iso, scaler, label_enc, features

st.sidebar.title("Model Loader")
MODEL_PATH = st.sidebar.text_input("Model path", MODEL_PATH_DEFAULT)
SCALER_PATH = st.sidebar.text_input("Scaler path", SCALER_PATH_DEFAULT)

try:
    classifier, iso_forest, scaler, label_encoder, trained_features = load_bundle(MODEL_PATH, SCALER_PATH)
    st.sidebar.success("Model & scaler loaded")
except Exception as e:
    st.sidebar.error(f"Failed to load model/scaler: {e}")
    st.stop()

     
# UI Inputs
     
st.title("CICIDS2018 Chunked Dashboard — CSV + Excel export")
uploaded = st.file_uploader("Upload CSV (or leave blank to use local path)", type=["csv"])
local_path = st.text_input("Or enter local CSV path (recommended for large files):")
chunk_size = st.number_input("Chunk size (rows)", value=DEFAULT_CHUNK_SIZE, min_value=2000, step=1000)
start = st.button("Start Processing")

# estimate total rows/chunks when local path provided for progress bar
def estimate_total_rows(path, max_check=5_000_000):
    try:
        with open(path, "rb") as fh:
            total = 0
            for i, _ in enumerate(fh, 1):
                total = i
                if i > max_check:
                    break
        return max(0, total - 1)
    except Exception:
        return None

if start and not st.session_state.processed:

    # validate data source
    if local_path:
        if not os.path.exists(local_path):
            st.error("Local path does not exist.")
            st.stop()
        csv_source = ("path", local_path)
        estimated_rows = estimate_total_rows(local_path)
        estimated_chunks = math.ceil(estimated_rows / chunk_size) if estimated_rows else None
    elif uploaded:
        csv_source = ("upload", uploaded)
        estimated_chunks = None
    else:
        st.error("Provide a CSV via upload or a local path.")
        st.stop()

    # accumulators
    predictions = []
    iso_scores = []
    true_labels = []
    timestamps = []
    probs_list = []

    label_col = None
    timestamp_col = None
    first_chunk = True
    chunk_idx = 0

    status = st.empty()
    prog = st.progress(0)

    # read in chunks
    try:
        if csv_source[0] == "path":
            reader = pd.read_csv(csv_source[1], chunksize=chunk_size, iterator=True, low_memory=True)
        else:
            upload_file = csv_source[1]
            upload_file.seek(0)
            reader = pd.read_csv(io.BytesIO(upload_file.getvalue()), chunksize=chunk_size, iterator=True, low_memory=True)
    except Exception as e:
        st.error(f"Failed to open CSV: {e}")
        st.stop()

    for chunk in reader:
        chunk_idx += 1
        status.info(f"Processing chunk #{chunk_idx} (rows ~ {len(chunk):,})")
        # normalize columns
        chunk = normalize_columns(chunk)

        # detect label & timestamp on first chunk
        if first_chunk:
            cols = list(chunk.columns)
            label_candidates = ["label", "attack_cat", "attack_type", "attack", "type", "flowlabel"]
            label_col = next((c for c in cols if c in label_candidates), None)
            ts_candidates = [c for c in cols if "time" in c or "timestamp" in c or "date" in c or c == "ts"]
            timestamp_col = ts_candidates[0] if ts_candidates else None
            first_chunk = False
            status.write(f"Detected label column: {label_col}  |  timestamp column: {timestamp_col}")

        # Remove accidental header rows embedded as data (e.g., rows where label == 'label' or empty)
        if label_col and label_col in chunk.columns:
            chunk = chunk[~chunk[label_col].astype(str).str.lower().isin({label_col, "label", "", "nan", "none"})]

        # preserve true labels/timestamps for this chunk
        if label_col and label_col in chunk.columns:
            true_part = chunk[label_col].astype(str).tolist()
        else:
            true_part = None

        if timestamp_col and timestamp_col in chunk.columns:
            ts_part = chunk[timestamp_col].tolist()
        else:
            ts_part = None

        # features -> numeric -> align
        feat_chunk = chunk.drop(columns=[c for c in (label_col, timestamp_col) if c in chunk.columns])
        feat_chunk = safe_to_numeric(feat_chunk)

        if trained_features:
            for f in trained_features:
                if f not in feat_chunk.columns:
                    feat_chunk[f] = 0
            feat_chunk = feat_chunk[trained_features]

        # scale and predict
        try:
            X_scaled = scaler.transform(feat_chunk)
        except Exception:
            feat_chunk = feat_chunk.replace([np.inf, -np.inf], np.nan).fillna(0)
            X_scaled = scaler.transform(feat_chunk)

        try:
            preds_enc = classifier.predict(X_scaled)
        except Exception as e:
            st.error(f"Classifier failed on chunk {chunk_idx}: {e}")
            preds_enc = np.array([None] * X_scaled.shape[0])

        if label_encoder is not None:
            try:
                preds_chunk = label_encoder.inverse_transform(preds_enc)
            except Exception:
                preds_chunk = preds_enc
        else:
            preds_chunk = preds_enc

        # probabilities
        chunk_probs = None
        try:
            chunk_probs = classifier.predict_proba(X_scaled)
        except Exception:
            try:
                dec = classifier.decision_function(X_scaled)
                if dec.ndim == 1:
                    probs0 = 1 / (1 + np.exp(dec))
                    chunk_probs = np.vstack([1-probs0, probs0]).T
                else:
                    exp = np.exp(dec - np.max(dec, axis=1, keepdims=True))
                    chunk_probs = exp / np.sum(exp, axis=1, keepdims=True)
            except Exception:
                chunk_probs = None

        # isolation forest scores
        if iso_forest is not None:
            try:
                iso_score_chunk = iso_forest.decision_function(X_scaled)
            except Exception:
                iso_score_chunk = np.zeros(X_scaled.shape[0])
        else:
            iso_score_chunk = np.zeros(X_scaled.shape[0])

        # append results
        predictions.extend(list(preds_chunk))
        iso_scores.extend(list(iso_score_chunk))
        if true_part:
            true_labels.extend(true_part)
        if ts_part:
            timestamps.extend(ts_part)
        if chunk_probs is not None:
            probs_list.append(chunk_probs)

        # update progress (estimate if possible)
        if estimated_chunks:
            prog.progress(min(1.0, chunk_idx / float(estimated_chunks)))
        else:
            # heuristic: advance a bit per chunk
            prog.progress(min(1.0, min(0.99, chunk_idx * 0.05)))

    status.success("All chunks processed.")

    #                ---
    # Ensure arrays have same length (pad if needed) to avoid malformed DataFrame
    #                ---
    risk_levels = [classify_risk(p, s) for p, s in zip(predictions, iso_scores)]

    # pad arrays
    lists = {
        "prediction": predictions,
        "iso_score": iso_scores,
        "risk": risk_levels,
        "true_label": true_labels,
        "timestamp": timestamps
    }
    max_len = max(len(v) for v in lists.values())

    def pad(lst, n, pad=None):
        lst = list(lst)
        if len(lst) < n:
            lst.extend([pad] * (n - len(lst)))
        return lst

    for k in lists:
        lists[k] = pad(lists[k], max_len, pad=None)

    out_df = pd.DataFrame({
        "prediction": lists["prediction"],
        "iso_score": lists["iso_score"],
        "risk_level": lists["risk"],
        "true_label": lists["true_label"],
        "timestamp": lists["timestamp"]
    })

    # combine probs safe
    prob_matrix = None
    if probs_list:
        try:
            prob_matrix = np.vstack(probs_list)
        except Exception:
            prob_matrix = None

    # classes
    classes = None
    if out_df["true_label"].notna().any():
        classes = sorted([str(x) for x in pd.unique(out_df["true_label"].dropna().astype(str).tolist() + out_df["prediction"].dropna().astype(str).tolist())])

    # timeline
    line_df = None
    if "timestamp" in out_df.columns and out_df["timestamp"].notna().any():
        try:
            out_df["__ts"] = pd.to_datetime(out_df["timestamp"], errors="coerce")
            out_df["__ts_floor"] = out_df["__ts"].dt.floor("min")
            line_df = out_df.groupby(["__ts_floor", "prediction"]).size().unstack(fill_value=0)
        except Exception:
            line_df = None

    # store in session_state
    st.session_state.processed = True
    st.session_state.out_df = out_df
    st.session_state.prob_matrix = prob_matrix
    st.session_state.classes = classes
    st.session_state.line_df = line_df

    st.success("Processing finished and stored in session state.")

  
# Display results (read-only from session)
  
if st.session_state.processed:
    out_df = st.session_state.out_df
    prob_matrix = st.session_state.prob_matrix
    classes = st.session_state.classes
    line_df = st.session_state.line_df

    st.header("Summary")
    st.markdown(f"- Rows processed: **{len(out_df):,}**")
    st.markdown(f"- Unique predicted labels: **{out_df['prediction'].nunique()}**")

    st.subheader("Prediction distribution (top labels)")
    pred_dist = out_df["prediction"].value_counts().reset_index().rename(columns={"index": "label", "prediction": "count_predictions"})
    st.dataframe(pred_dist.head(50))

    # risk distribution with color labels
    st.subheader("Risk Level Distribution")
    risk_counts = out_df["risk_level"].value_counts().reindex(["HIGH","MEDIUM","LOW"]).fillna(0)
    fig_risk, ax_risk = plt.subplots(figsize=(5,3))
    color_map = {"HIGH":"#FF0000","MEDIUM":"#FFA500","LOW":"#00AA00"}
    bars = ax_risk.bar(risk_counts.index, risk_counts.values, color=[color_map.get(x,"#333333") for x in risk_counts.index])
    for i, v in enumerate(risk_counts.values):
        ax_risk.text(i, v + max(1, int(0.01*max(risk_counts.values))), f"{int(v):,}", ha="center", va="bottom")
    ax_risk.set_title("Risk Level Distribution")
    st.pyplot(fig_risk)

    # confusion matrix
    if classes:
        st.header("Confusion Matrix")
        fig_cm, cm = plot_confusion_matrix_readable(out_df["true_label"].astype(str), out_df["prediction"].astype(str), classes, figsize=(max(8, len(classes)*0.3), max(6, len(classes)*0.3)))
        st.pyplot(fig_cm)

        st.header("Classification report (text)")
        st.text(classification_report(out_df["true_label"].astype(str), out_df["prediction"].astype(str), zero_division=0))
    else:
        st.info("True labels not found; confusion matrix and classification report are unavailable.")

    # ROC / PR if available
    if classes and prob_matrix is not None:
        try:
            st.header("ROC Curves")
            # user may implement plot_roc_multiclass if required; omitted for brevity
        except Exception:
            pass

    # anomaly histogram
    st.header("Anomaly score distribution")
    fig_iso, ax_iso = plt.subplots(figsize=(8,4))
    ax_iso.hist(out_df["iso_score"].dropna().astype(float), bins=80)
    ax_iso.set_title("IsolationForest score distribution")
    st.pyplot(fig_iso)

    # timeline
    if line_df is not None:
        st.header("Attack timeline (per minute)")
        st.line_chart(line_df)
    else:
        st.info("No usable timestamps for timeline.")

      
    # CSV download (safe)
      
    st.header("Download predictions CSV")
    csv_bytes = out_df.to_csv(index=False).encode("utf-8")
    st.download_button("Download CSV", csv_bytes, file_name="cicids_predictions.csv", mime="text/csv")

      
    # EXCEL Export
      
    st.header("Download Excel (Predictions + Pivot + Risk Chart + Confusion Matrix)")

    if st.button("Generate Excel file"):

        wb = Workbook()

        # Sheet 1: Predictions (write dataframe correctly with headers horizontally)
        ws1 = wb.active
        ws1.title = "Predictions"
        for r in dataframe_to_rows(out_df, index=False, header=True):
            ws1.append(r)
        for cell in ws1[1]:
            cell.font = Font(bold=True)

        # Sheet 2: Pivot
        ws2 = wb.create_sheet("Pivot")
        if "true_label" in out_df.columns and out_df["true_label"].notna().any():
            pivot = pd.pivot_table(out_df, index="true_label", columns="prediction", values="iso_score", aggfunc="count", fill_value=0)
            # write header + rows
            ws2.append(["true_label"] + list(pivot.columns))
            for idx, row in pivot.iterrows():
                ws2.append([idx] + list(row))
            for cell in ws2[1]:
                cell.font = Font(bold=True)
        else:
            ws2.append(["No true labels present; pivot table requires true labels."])

        # Sheet 3: Risk Chart (counts)
        ws3 = wb.create_sheet("Risk Chart")
        ws3.append(["risk", "count"])
        # ensure order HIGH, MEDIUM, LOW
        for key in ["HIGH", "MEDIUM", "LOW"]:
            ws3.append([key, int(risk_counts.get(key, 0))])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        # create bar chart
        chart = BarChart()
        chart.title = "Risk Level Distribution"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Risk"
        data = Reference(ws3, min_col=2, min_row=1, max_row=4)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws3.add_chart(chart, "E2")

        # Sheet 4: Confusion Matrix (raw numbers table)
        ws4 = wb.create_sheet("Confusion Matrix")
        if classes:
            cm = confusion_matrix(out_df["true_label"].astype(str), out_df["prediction"].astype(str), labels=classes)
            ws4.append([""] + classes)
            for idx, row in enumerate(cm):
                ws4.append([classes[idx]] + list(row))
            for cell in ws4[1]:
                cell.font = Font(bold=True)
        else:
            ws4.append(["No true labels => confusion matrix not available."])

        # Save to buffer and offer download
        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)
        st.download_button("Download Excel", excel_buf, file_name="cicids_report.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("Processing not started. Provide CSV (upload or path) and click 'Start Processing'.")
